import re
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from sklearn.ensemble import IsolationForest
from openpyxl import load_workbook


def count_digits(p):
    return sum(c.isdecimal() for c in p)


def count_lower(p):
    return sum(c.islower() for c in p)


def count_upper(p):
    return sum(c.isupper() for c in p)


def split_group(p):
    return re.findall('\d+|[A-Z]+|[a-z]+|[^A-Za-z\d]+', p)


def unique(p):
    return len(set(p))


def seq(p):
    v = 1
    for i in range(1,len(p)):
        if abs(ord(p[i]) - ord(p[i-1])) > 1:
            v += 1
    return v


def convert_to_dataframe(file):
    f = open(file, encoding="utf-8")
    data = [x.strip('\n') for x in f.readlines()]
    f.close()
    df = pd.DataFrame(data, columns=['password'])

    if m1_length:
        df['length'] = df.apply(lambda row: len(row.password), axis=1)
    if m2_lowercase:
        df['lower'] = df.apply(lambda row: count_lower(row.password), axis=1)
    if m3_uppercase:
        df['upper'] = df.apply(lambda row: count_upper(row.password), axis=1)
    if m4_digits:
        df['digits'] = df.apply(lambda row: count_digits(row.password), axis=1)
    if m5_symbols:
        df['symbols'] = df.apply(lambda row: len(row.password) - row.lower - row.upper - row.digits, axis=1)
    if m6_groups:
        df['groups'] = df.apply(lambda row: len(split_group(row.password)), axis=1)
    if m7_unique:
        df['unique'] = df.apply(lambda row: unique(row.password), axis=1)
    if m8_seq:
        df['seq'] = df.apply(lambda row: seq(row.password), axis=1)
    return df


def train_model(trainfile):
    # random_state = np.random.RandomState(42)
    model = IsolationForest(
        n_estimators=num_trees,
        max_features=num_features,
        max_samples=num_samples,
        verbose=0)
    model.fit(convert_to_dataframe(trainfile).drop('password', axis=1))
    return model


def calculate_score(s):
    return s*(-100000)


def score(model, testfile):
    data = convert_to_dataframe(testfile)
    scores = model.score_samples(data.drop('password', axis=1))
    data['score'] = scores
    data['if'] = data.apply(lambda row: calculate_score(row.score), axis=1)
    return data


def save_to_excel(data, file):
    wb = load_workbook(file)
    ws = wb['Sheet1']

    for index, row in data.iterrows():
        cell = 'C%d' % (index + 2)
        ws[cell] = row['score']

    for index, row in data.iterrows():
        cell = 'D%d' % (index + 2)
        ws[cell] = row['if']

    wb.save(file)


def check_sequence_string(s):
    abc = 'abcdefghijklmnopqrstuvwxyz'
    keyboard = 'qwertyuiopasdfghjklzxcvbnm'
    numbers = '0123456789'
    numbers_odd = '13579'
    numbers_even = '02468'
    numpad = '0147258369'

    s_r = s[::-1]

    if s in abc or s in keyboard or s in numbers or s in numpad or s in numbers_odd or s in numbers_even \
            or s_r in abc or s_r in keyboard or s_r in numbers or s_r in numpad or s_r in numbers_odd \
            or s_r in numbers_even:
        return True
    else:
        return False


def check_sequence_password(p):
    const = 0.4
    if len(set(p)) == 1:
        return const

    if (check_sequence_string(p[::2]) or len(set(p[::2])) == 1) \
            and (check_sequence_string(p[1::2]) or len(set(p[1::2])) == 1):
        return const

    seq = '0' * len(p)

    for i in range(2, len(p)):
        for j in range(0, i - 1):
            if not seq[j:i + 1] == (i + 1 - j) * '1':
                if check_sequence_string(p[j:i + 1]):
                    seq = seq[:j] + (i + 1 - j) * '1' + seq[i + 1:]

    res = seq.count('1') / len(p)
    return const + (1-const) * (1 - res)


def adjust_scores(data):
    data['if'] = data.apply(lambda row: check_sequence_password(row.password)*row['if'], axis=1)
    return data


def run():
    model = train_model(train_file)
    testdata1_scored = score(model, test1_file)
    testdata1_scored = adjust_scores(testdata1_scored)
    save_to_excel(testdata1_scored, test1_fileX)
    testdata2_scored = score(model, testDS2)
    testdata2_scored = adjust_scores(testdata2_scored)
    save_to_excel(testdata2_scored, testDS2X)


dataset_path = 'datasets/clean/'
trainDS1 = dataset_path + 'train/train_set1.txt'    #10k most common
trainDS2 = dataset_path + 'train/train_set2.txt'    #30k CS under 10**6
trainDS3 = dataset_path + 'train/train_set3.txt'    #30k RY random sample
trainDS4 = dataset_path + 'train/train_set4.txt'    #30k CS over 10**14
trainDS5 = dataset_path + 'train/train_set5.txt'    #10k generated

testDS11 = dataset_path + 'test/test_set11.txt'
testDS11X = dataset_path + 'test/test_set11.xlsx'
testDS12 = dataset_path + 'test/test_set12.txt'
testDS12X = dataset_path + 'test/test_set12.xlsx'

testDS2 = dataset_path + 'test/test_set2.txt'
testDS2X = dataset_path + 'test/test_set2.xlsx'

train_file = trainDS1
test1_file = testDS12
test1_fileX = testDS12X

m1_length = True
m2_lowercase = True
m3_uppercase = True
m4_digits = True
m5_symbols = True
m6_groups = True
m7_unique = True
m8_seq = True

#Hyperparameters
num_trees = 100
num_features = 8
num_samples = 'auto'

###########
run()
